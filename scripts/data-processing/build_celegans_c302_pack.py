#!/usr/bin/env python3
"""Build a checksummed DFLY v1 pack from the OpenWorm c302 C. elegans workbook.

The source workbook must be obtained outside this repository. This converter keeps a compact
neuron-to-neuron graph only: `NMJ` rows are counted as source neuromuscular evidence but are not
silently treated as neuronal edges. Source connectivity remains `SOURCE DATA`; the browser's
sensors, motor decoder, body and any generated spatial layout remain `MODELLED MAPPING`.
"""

from __future__ import annotations

import argparse
import hashlib
import json
import re
import shutil
import sys
import tempfile
from dataclasses import dataclass
from pathlib import Path
from typing import Any, Iterator

import numpy as np
from openpyxl import load_workbook


PROJECT_ROOT = Path(__file__).resolve().parents[2]
TRANSFORM_NAME = "c-elegans-c302-xlsx-to-dfly"
TRANSFORM_VERSION = "1"
REQUIRED_HEADERS = ("Neuron 1", "Neuron 2", "Type", "Nbr")
CHEMICAL_TYPES = frozenset(("R", "Rp", "S", "Sp"))
ELECTRICAL_TYPE = "EJ"
NEUROMUSCULAR_TYPE = "NMJ"


@dataclass(frozen=True)
class InputProvenance:
    source_url: str
    license_name: str
    citations: tuple[str, ...]


@dataclass(frozen=True)
class ColumnDefinition:
    name: str
    scalar_type: str
    element_count: int
    semantic_status: str = "SOURCE DATA"
    stride: int = 1


@dataclass(frozen=True)
class ConnectionRecord:
    source: str
    target: str
    weight: float
    connection_kind: int


def sha256_file(path: Path) -> str:
    digest = hashlib.sha256()
    with path.open("rb") as handle:
        for block in iter(lambda: handle.read(8 * 1024 * 1024), b""):
            digest.update(block)
    return digest.hexdigest()


def is_within(path: Path, parent: Path) -> bool:
    try:
        path.relative_to(parent)
        return True
    except ValueError:
        return False


def normalise_cell_name(value: object) -> str | None:
    name = "" if value is None else str(value).strip()
    if not name or name == "NMJ" or not name[0].isupper():
        return None
    return re.sub(r"0(\d)$", r"\1", name)


def validate_headers(path: Path) -> None:
    workbook = load_workbook(path, read_only=True, data_only=True)
    worksheet = workbook.active
    header = tuple("" if value is None else str(value).strip() for value in next(worksheet.iter_rows(max_row=1, values_only=True)))
    missing = [name for name in REQUIRED_HEADERS if name not in header]
    if missing:
        raise ValueError(f"Workbook is missing required columns: {', '.join(missing)}")


def iter_records(path: Path) -> Iterator[tuple[str, ConnectionRecord | None]]:
    workbook = load_workbook(path, read_only=True, data_only=True)
    worksheet = workbook.active
    header = ["" if value is None else str(value).strip() for value in next(worksheet.iter_rows(max_row=1, values_only=True))]
    indices = {name: header.index(name) for name in REQUIRED_HEADERS}
    for row_number, row in enumerate(worksheet.iter_rows(min_row=2, values_only=True), start=2):
        source = normalise_cell_name(row[indices["Neuron 1"]])
        target = normalise_cell_name(row[indices["Neuron 2"]])
        connection_type = "" if row[indices["Type"]] is None else str(row[indices["Type"]]).strip()
        raw_weight = row[indices["Nbr"]]
        if connection_type == NEUROMUSCULAR_TYPE:
            yield "NMJ", None
            continue
        if connection_type not in CHEMICAL_TYPES and connection_type != ELECTRICAL_TYPE:
            raise ValueError(f"Unsupported connection type {connection_type!r} on row {row_number}.")
        if source is None or target is None:
            raise ValueError(f"Neuron-to-neuron row {row_number} includes an invalid cell name.")
        try:
            weight = float(raw_weight)
        except (TypeError, ValueError) as error:
            raise ValueError(f"Row {row_number} has an invalid Nbr weight: {raw_weight!r}") from error
        if not np.isfinite(weight) or weight < 0:
            raise ValueError(f"Row {row_number} has a negative or non-finite weight: {raw_weight!r}")
        if weight == 0:
            yield "ZERO", None
            continue
        yield "CONNECTION", ConnectionRecord(source, target, weight, 1 if connection_type == ELECTRICAL_TYPE else 0)


def prepare_output_dir(path: Path) -> None:
    if path.exists() and any(path.iterdir()):
        raise ValueError(f"Output directory must be empty: {path}")
    path.mkdir(parents=True, exist_ok=True)


def chunk_column(source: Path, output_dir: Path, column: ColumnDefinition, chunk_bytes: int) -> list[dict[str, Any]]:
    scalar_dtypes = {"u16": "<u2", "u32": "<u4", "f32": "<f4"}
    if column.scalar_type not in scalar_dtypes:
        raise ValueError(f"Unsupported column scalar type: {column.scalar_type}")
    item_bytes = np.dtype(scalar_dtypes[column.scalar_type]).itemsize * column.stride
    chunks_dir = output_dir / "chunks"
    chunks_dir.mkdir(exist_ok=True)
    chunks: list[dict[str, Any]] = []
    item_offset = 0
    with source.open("rb") as input_handle:
        part = 0
        while block := input_handle.read(chunk_bytes):
            if len(block) % item_bytes:
                raise ValueError(f"Column {column.name} is not aligned to its declared element stride.")
            filename = f"{column.name}-{part:05d}.bin"
            target = chunks_dir / filename
            target.write_bytes(block)
            item_count = len(block) // item_bytes
            chunks.append(
                {
                    "id": f"{column.name}:{part}",
                    "column": column.name,
                    "path": f"chunks/{filename}",
                    "bytes": len(block),
                    "sha256": sha256_file(target),
                    "elementOffset": item_offset,
                    "elementCount": item_count,
                }
            )
            item_offset += item_count
            part += 1
    if item_offset != column.element_count:
        raise ValueError(f"Column {column.name} wrote {item_offset} values; expected {column.element_count}.")
    return chunks


def build_pack(
    workbook_path: Path,
    output_dir: Path,
    provenance: InputProvenance,
    chunk_mib: int = 4,
) -> dict[str, Any]:
    validate_headers(workbook_path)
    neuron_names: set[str] = set()
    records: list[ConnectionRecord] = []
    excluded_nmj = 0
    excluded_zero_weight = 0
    for record_kind, record in iter_records(workbook_path):
        if record_kind == "NMJ":
            excluded_nmj += 1
            continue
        if record_kind == "ZERO":
            excluded_zero_weight += 1
            continue
        if record is None:
            raise ValueError("A connection record was unexpectedly empty.")
        neuron_names.add(record.source)
        neuron_names.add(record.target)
        records.append(record)
    if not records or not neuron_names:
        raise ValueError("The source workbook produced no valid neuron-to-neuron connections.")
    if len(records) >= 2**32:
        raise ValueError("DFLY v1 uses Uint32 edge offsets and cannot encode this many edges.")

    cells = sorted(neuron_names)
    cell_to_index = {cell: index for index, cell in enumerate(cells)}
    incoming_counts = np.zeros(len(cells), dtype=np.uint64)
    for record in records:
        incoming_counts[cell_to_index[record.target]] += 1
    incoming_offsets = np.zeros(len(cells) + 1, dtype=np.uint32)
    np.cumsum(incoming_counts, dtype=np.uint64, out=incoming_offsets[1:])
    if int(incoming_offsets[-1]) != len(records):
        raise ValueError("CSR offsets are inconsistent with the accepted connection count.")

    prepare_output_dir(output_dir)
    with tempfile.TemporaryDirectory(prefix="dfly-celegans-") as temp_dir_string:
        temp_dir = Path(temp_dir_string)
        cell_indices_file = temp_dir / "cell-index.u32.bin"
        offsets_file = temp_dir / "incoming-offsets.u32.bin"
        source_file = temp_dir / "source-index.u32.bin"
        weight_file = temp_dir / "connection-weight.f32.bin"
        kind_file = temp_dir / "connection-kind.u16.bin"
        np.arange(len(cells), dtype="<u4").tofile(cell_indices_file)
        incoming_offsets.astype("<u4", copy=False).tofile(offsets_file)
        source_indices = np.memmap(source_file, mode="w+", dtype="<u4", shape=(len(records),))
        weights = np.memmap(weight_file, mode="w+", dtype="<f4", shape=(len(records),))
        kinds = np.memmap(kind_file, mode="w+", dtype="<u2", shape=(len(records),))
        cursors = incoming_offsets[:-1].copy()
        for record in records:
            target_index = cell_to_index[record.target]
            slot = int(cursors[target_index])
            source_indices[slot] = cell_to_index[record.source]
            weights[slot] = record.weight
            kinds[slot] = record.connection_kind
            cursors[target_index] += 1
        if not np.array_equal(cursors, incoming_offsets[1:]):
            raise ValueError("CSR writes did not fill every target interval.")
        source_indices.flush()
        weights.flush()
        kinds.flush()

        definitions = (
            ColumnDefinition("cell_index", "u32", len(cells)),
            ColumnDefinition("incoming_offsets", "u32", len(cells) + 1),
            ColumnDefinition("source_index", "u32", len(records)),
            ColumnDefinition("connection_weight", "f32", len(records)),
            ColumnDefinition("connection_kind", "u16", len(records)),
        )
        paths = {
            "cell_index": cell_indices_file,
            "incoming_offsets": offsets_file,
            "source_index": source_file,
            "connection_weight": weight_file,
            "connection_kind": kind_file,
        }
        all_chunks: list[dict[str, Any]] = []
        columns_manifest: dict[str, Any] = {}
        for definition in definitions:
            chunks = chunk_column(paths[definition.name], output_dir, definition, chunk_mib * 1024 * 1024)
            all_chunks.extend(chunks)
            columns_manifest[definition.name] = {
                "scalarType": definition.scalar_type,
                "elementCount": definition.element_count,
                "stride": definition.stride,
                "semanticStatus": definition.semantic_status,
                "chunks": [chunk["id"] for chunk in chunks],
            }

    manifest = {
        "format": "DFLY",
        "formatVersion": 1,
        "datasetId": "celegans-c302-hermaphrodite-connectivity",
        "release": "c302-master-6cd861f8ca4d3241ee9cf4627884caa930dab53c",
        "origin": provenance.source_url,
        "license": provenance.license_name,
        "neuronCount": len(cells),
        "synapseCount": len(records),
        "provenance": {
            "sourceFiles": [{"name": workbook_path.name, "sha256": sha256_file(workbook_path)}],
            "citations": list(provenance.citations),
            "transform": {
                "name": TRANSFORM_NAME,
                "version": TRANSFORM_VERSION,
                "excludedNeuromuscularRows": excluded_nmj,
                "excludedZeroWeightRows": excluded_zero_weight,
                "connectionKindOrder": ["chemical", "electrical"],
            },
        },
        "dictionaries": {"cellNames": cells, "connectionKinds": ["chemical", "electrical"]},
        "columns": columns_manifest,
        "chunks": all_chunks,
    }
    (output_dir / "manifest.json").write_text(json.dumps(manifest, indent=2, sort_keys=True) + "\n", encoding="utf-8")
    return manifest


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument("--workbook", type=Path, required=True, help="User-downloaded c302 NeuronConnectFormatted.xlsx")
    parser.add_argument("--out-dir", type=Path, required=True, help="Empty directory outside this Git repository")
    parser.add_argument("--source-url", default="https://github.com/openworm/c302/blob/6cd861f8ca4d3241ee9cf4627884caa930dab53c/c302/data/NeuronConnectFormatted.xlsx")
    parser.add_argument("--license", dest="license_name", default="MIT")
    parser.add_argument("--citation", action="append", default=[], help="Repeat for every source citation")
    parser.add_argument("--chunk-mib", type=int, default=4, help="Maximum binary chunk size in MiB")
    parser.add_argument("--accept-source-terms", action="store_true", help="Required acknowledgement before converting the user-downloaded source")
    return parser.parse_args()


def main() -> int:
    args = parse_args()
    if not args.accept_source_terms:
        raise SystemExit("Refusing conversion: supply --accept-source-terms after reviewing the c302 MIT licence and upstream citations.")
    if args.chunk_mib < 1 or args.chunk_mib > 1024:
        raise SystemExit("--chunk-mib must be between 1 and 1024.")
    workbook_path = args.workbook.resolve()
    output_dir = args.out_dir.resolve()
    if is_within(output_dir, PROJECT_ROOT):
        raise SystemExit("Refusing output inside the repository. Choose external storage such as /home/ubuntu/celegans-packs/c302.")
    if not workbook_path.is_file():
        raise SystemExit(f"Workbook does not exist: {workbook_path}")
    citations = tuple(args.citation) or (
        "Gleeson et al. (2018), c302: a multiscale framework for modelling the nervous system of Caenorhabditis elegans.",
        "Varshney et al. (2011), Structural properties of the Caenorhabditis elegans neuronal network.",
    )
    try:
        manifest = build_pack(workbook_path, output_dir, InputProvenance(args.source_url, args.license_name, citations), args.chunk_mib)
    except (OSError, ValueError) as error:
        shutil.rmtree(output_dir, ignore_errors=True)
        raise SystemExit(f"DFLY conversion failed: {error}") from error
    print(json.dumps({"manifest": str(output_dir / "manifest.json"), "neurons": manifest["neuronCount"], "edges": manifest["synapseCount"], "chunks": len(manifest["chunks"])}, indent=2))
    return 0


if __name__ == "__main__":
    sys.exit(main())
