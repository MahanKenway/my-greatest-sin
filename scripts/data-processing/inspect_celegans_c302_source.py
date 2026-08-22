#!/usr/bin/env python3
"""Read-only schema inspector for a user-downloaded c302 C. elegans workbook.

This utility only prints sheet titles, dimensions, and the first non-empty rows. It never copies
source data into the repository or transforms it into a distributable artefact.
"""

from __future__ import annotations

import argparse
from pathlib import Path

from openpyxl import load_workbook


def main() -> int:
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument("workbook", type=Path)
    args = parser.parse_args()
    workbook_path = args.workbook.resolve()
    if not workbook_path.is_file():
        raise SystemExit(f"Workbook does not exist: {workbook_path}")
    workbook = load_workbook(workbook_path, read_only=True, data_only=True)
    for worksheet in workbook.worksheets:
        print(f"[{worksheet.title}] rows={worksheet.max_row} columns={worksheet.max_column}")
        shown = 0
        nodes: set[str] = set()
        types: dict[str, int] = {}
        nmj_examples: list[tuple[str, str]] = []
        for row in worksheet.iter_rows(values_only=True):
            if not any(value is not None and str(value).strip() for value in row):
                continue
            if shown > 0 and len(row) >= 4:
                left = "" if row[0] is None else str(row[0]).strip()
                right = "" if row[1] is None else str(row[1]).strip()
                connection_type = "" if row[2] is None else str(row[2]).strip()
                nodes.update(node for node in (left, right) if node)
                types[connection_type] = types.get(connection_type, 0) + 1
                if connection_type == "NMJ" and len(nmj_examples) < 12:
                    nmj_examples.append((left, right))
            print(" | ".join("" if value is None else str(value).strip() for value in row[:12]))
            shown += 1
            if shown == 8:
                break
        if worksheet.max_row > shown:
            for row in list(worksheet.iter_rows(min_row=shown + 1, values_only=True)):
                if len(row) < 4:
                    continue
                left = "" if row[0] is None else str(row[0]).strip()
                right = "" if row[1] is None else str(row[1]).strip()
                connection_type = "" if row[2] is None else str(row[2]).strip()
                nodes.update(node for node in (left, right) if node)
                types[connection_type] = types.get(connection_type, 0) + 1
                if connection_type == "NMJ" and len(nmj_examples) < 12:
                    nmj_examples.append((left, right))
        if nodes:
            print(f"unique_nodes={len(nodes)}")
            print("edge_types=" + ", ".join(f"{name}:{count}" for name, count in sorted(types.items())))
            non_neuronal = sorted(node for node in nodes if node.startswith(("dBWM", "vBWM", "pm", "vm", "um")))
            print("non_neuronal_prefix_nodes=" + ",".join(non_neuronal[:20]))
            print("nmj_examples=" + ", ".join(f"{left}->{right}" for left, right in nmj_examples))
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
